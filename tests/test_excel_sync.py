"""Tests for excel_sync.py — reading the two Power Automate result workbooks."""
import datetime

import openpyxl
import pytest

import excel_sync


def _write_oc_contacts(path, rows, header_row=1):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "OrderContacts"
    for _ in range(header_row - 1):
        ws.append([])
    ws.append(["Folder_Name", "Logistics_Coordinator", "Logistics_Coordinator_Email", "RSM", "RSM_Email"])
    for row in rows:
        ws.append(row)
    wb.save(path)


def _write_shipping_dates(path, rows, header_row=7):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ShippingDates"
    for _ in range(header_row - 1):
        ws.append([])
    ws.append(["Folder_Name", "Shipping_Date", "Reason", "Document"])
    for row in rows:
        ws.append(row)
    wb.save(path)


# ── lookup_oc_contacts ──────────────────────────────────────────────────────

def test_lookup_oc_contacts_returns_mapped_fields(tmp_path):
    path = tmp_path / "OC_Contacts.xlsx"
    _write_oc_contacts(path, [
        ["DO001 ACME", "Jane Doe", "jane@example.com", "John Smith", "john@example.com"],
    ])
    result = excel_sync.lookup_oc_contacts(str(tmp_path / "DO001 ACME"), path=str(path))
    assert result == {
        "logistics_coordinator": "Jane Doe",
        "logistics_coordinator_email": "jane@example.com",
        "rsm": "John Smith",
        "rsm_email": "john@example.com",
    }


def test_lookup_oc_contacts_matches_case_insensitively_and_ignores_separator(tmp_path):
    path = tmp_path / "OC_Contacts.xlsx"
    _write_oc_contacts(path, [
        ["do001 acme", "Jane Doe", "", "", ""],
    ])
    result = excel_sync.lookup_oc_contacts(str(tmp_path / "DO001 ACME") + "\\", path=str(path))
    assert result == {"logistics_coordinator": "Jane Doe"}


def test_lookup_oc_contacts_last_matching_row_wins(tmp_path):
    path = tmp_path / "OC_Contacts.xlsx"
    _write_oc_contacts(path, [
        ["DO001 ACME", "Old Name", "", "", ""],
        ["DO001 ACME", "New Name", "", "", ""],
    ])
    result = excel_sync.lookup_oc_contacts(str(tmp_path / "DO001 ACME"), path=str(path))
    assert result == {"logistics_coordinator": "New Name"}


def test_lookup_oc_contacts_missing_file_returns_empty_dict(tmp_path):
    result = excel_sync.lookup_oc_contacts(str(tmp_path / "DO001 ACME"), path=str(tmp_path / "missing.xlsx"))
    assert result == {}


def test_lookup_oc_contacts_no_match_returns_empty_dict(tmp_path):
    path = tmp_path / "OC_Contacts.xlsx"
    _write_oc_contacts(path, [["DO999 Other", "Jane Doe", "", "", ""]])
    result = excel_sync.lookup_oc_contacts(str(tmp_path / "DO001 ACME"), path=str(path))
    assert result == {}


# ── lookup_shipping_date ─────────────────────────────────────────────────────

def test_lookup_shipping_date_formats_datetime_cell(tmp_path):
    path = tmp_path / "Dossier_Shipping_Date.xlsx"
    _write_shipping_dates(path, [
        ["DO001 ACME", datetime.datetime(2026, 1, 29), "Found on the invoice", "Invoice.pdf"],
    ])
    result = excel_sync.lookup_shipping_date(str(tmp_path / "DO001 ACME"), path=str(path))
    assert result == {
        "shipping_date": "1/29/2026",
        "reasoning": "Found on the invoice",
        "source_document": "Invoice.pdf",
    }


def test_lookup_shipping_date_returns_empty_when_no_match(tmp_path):
    path = tmp_path / "Dossier_Shipping_Date.xlsx"
    _write_shipping_dates(path, [
        ["DO999 Other", datetime.datetime(2026, 1, 29), "", ""],
    ])
    result = excel_sync.lookup_shipping_date(str(tmp_path / "DO001 ACME"), path=str(path))
    assert result == {}


def test_lookup_shipping_date_returns_empty_when_date_cell_blank(tmp_path):
    path = tmp_path / "Dossier_Shipping_Date.xlsx"
    _write_shipping_dates(path, [
        ["DO001 ACME", None, "", ""],
    ])
    result = excel_sync.lookup_shipping_date(str(tmp_path / "DO001 ACME"), path=str(path))
    assert result == {}


def test_lookup_shipping_date_returns_reason_when_date_blank_but_reason_present(tmp_path):
    path = tmp_path / "Dossier_Shipping_Date.xlsx"
    _write_shipping_dates(path, [
        ["DO001 ACME", None, "No shipping documents uploaded yet", ""],
    ])
    result = excel_sync.lookup_shipping_date(str(tmp_path / "DO001 ACME"), path=str(path))
    assert result == {
        "shipping_date": None,
        "reasoning": "No shipping documents uploaded yet",
        "source_document": None,
    }


def test_lookup_shipping_date_missing_file_returns_empty_dict(tmp_path):
    result = excel_sync.lookup_shipping_date(str(tmp_path / "DO001 ACME"), path=str(tmp_path / "missing.xlsx"))
    assert result == {}


def test_lookup_latest_shipping_result_handles_generic_subfolder_name(tmp_path):
    path = tmp_path / "Dossier_Shipping_Date.xlsx"
    _write_shipping_dates(path, [
        ["DO999 Other", datetime.datetime(2026, 1, 29), "Older result", "Old.pdf"],
        [
            "Shipping Documents and Invoices",
            None,
            "The files only contain freight quotes, not a confirmed shipping date.",
            None,
        ],
    ])

    result = excel_sync.lookup_latest_shipping_result(path=str(path))

    assert result == {
        "shipping_date": None,
        "reasoning": "The files only contain freight quotes, not a confirmed shipping date.",
        "source_document": None,
    }


def test_lookup_latest_shipping_result_never_borrows_another_orders_row(tmp_path):
    """A row naming a different order must not be stamped onto this order."""
    path = tmp_path / "Dossier_Shipping_Date.xlsx"
    _write_shipping_dates(path, [
        ["Shipping Documents and Invoices", None, "No shipping documents yet", None],
        ["DO126384 Ferrero OPP435046", datetime.datetime(2026, 6, 30), "FedEx POD", "POD.pdf"],
    ])

    result = excel_sync.lookup_latest_shipping_result(
        str(tmp_path / "DOXX SSAB OPP426584"), path=str(path)
    )

    assert result["shipping_date"] is None
    assert result["reasoning"] == "No shipping documents yet"


def test_lookup_latest_shipping_result_accepts_own_order_row(tmp_path):
    path = tmp_path / "Dossier_Shipping_Date.xlsx"
    _write_shipping_dates(path, [
        ["DO126384 Ferrero OPP435046", datetime.datetime(2026, 6, 30), "FedEx POD", "POD.pdf"],
    ])

    result = excel_sync.lookup_latest_shipping_result(
        str(tmp_path / "DO126384 Ferrero OPP435046"), path=str(path)
    )

    assert result["shipping_date"] == "6/30/2026"


def test_lookup_latest_shipping_result_without_folder_ignores_order_rows(tmp_path):
    """Called without a folder, only generic rows are usable."""
    path = tmp_path / "Dossier_Shipping_Date.xlsx"
    _write_shipping_dates(path, [
        ["DO126384 Ferrero OPP435046", datetime.datetime(2026, 6, 30), "FedEx POD", "POD.pdf"],
    ])

    assert excel_sync.lookup_latest_shipping_result(path=str(path)) == {}


# ── locked workbooks ────────────────────────────────────────────────────────

def test_read_rows_raises_workbook_unavailable_when_locked(tmp_path, monkeypatch):
    path = tmp_path / "OC_Contacts.xlsx"
    _write_oc_contacts(path, [["DO001 ACME", "Jane", "jane@x.com", "John", "john@x.com"]])

    def _locked(*args, **kwargs):
        raise PermissionError(13, "Permission denied")

    monkeypatch.setattr(excel_sync.openpyxl, "load_workbook", _locked)
    monkeypatch.setattr(excel_sync.time, "sleep", lambda s: None)

    with pytest.raises(excel_sync.WorkbookUnavailable, match="locked"):
        excel_sync._read_rows(str(path))


def test_read_rows_retries_and_succeeds_once_lock_clears(tmp_path, monkeypatch):
    path = tmp_path / "OC_Contacts.xlsx"
    _write_oc_contacts(path, [["DO001 ACME", "Jane", "jane@x.com", "John", "john@x.com"]])

    real_load = excel_sync.openpyxl.load_workbook
    calls = {"n": 0}

    def _flaky(*args, **kwargs):
        calls["n"] += 1
        if calls["n"] == 1:
            raise PermissionError(13, "Permission denied")
        return real_load(*args, **kwargs)

    monkeypatch.setattr(excel_sync.openpyxl, "load_workbook", _flaky)
    monkeypatch.setattr(excel_sync.time, "sleep", lambda s: None)

    rows = excel_sync._read_rows(str(path))

    assert calls["n"] == 2
    assert rows[0]["Folder_Name"] == "DO001 ACME"


def test_lookup_oc_contacts_propagates_locked_workbook(tmp_path, monkeypatch):
    path = tmp_path / "OC_Contacts.xlsx"
    _write_oc_contacts(path, [["DO001 ACME", "Jane", "jane@x.com", "John", "john@x.com"]])

    monkeypatch.setattr(
        excel_sync.openpyxl, "load_workbook",
        lambda *a, **k: (_ for _ in ()).throw(PermissionError(13, "Permission denied")),
    )
    monkeypatch.setattr(excel_sync.time, "sleep", lambda s: None)

    with pytest.raises(excel_sync.WorkbookUnavailable):
        excel_sync.lookup_oc_contacts(str(tmp_path / "DO001 ACME"), path=str(path))


def test_lookup_paths_follow_config_without_reimport(tmp_path, monkeypatch):
    """Default workbook paths must resolve per call, not at import time."""
    path = tmp_path / "OC_Contacts.xlsx"
    _write_oc_contacts(path, [["DO001 ACME", "Jane", "jane@x.com", "John", "john@x.com"]])
    monkeypatch.setattr(
        "webapp.backend.settings.load_config", lambda: {"excel_root": str(tmp_path)}
    )

    result = excel_sync.lookup_oc_contacts(str(tmp_path / "DO001 ACME"))

    assert result["logistics_coordinator"] == "Jane"


# ── waiting for the flow's row to sync down ─────────────────────────────────

def test_wait_for_oc_contacts_retries_until_row_appears(tmp_path, monkeypatch):
    path = tmp_path / "OC_Contacts.xlsx"
    _write_oc_contacts(path, [])
    monkeypatch.setattr(excel_sync.time, "sleep", lambda s: None)

    calls = {"n": 0}
    real = excel_sync.lookup_oc_contacts

    def _late(folder, path=None):
        calls["n"] += 1
        if calls["n"] == 3:
            _write_oc_contacts(path, [["DO001 ACME", "Jane", "j@x.com", "John", "jo@x.com"]])
        return real(folder, path=path)

    monkeypatch.setattr(excel_sync, "lookup_oc_contacts", _late)

    result = excel_sync.wait_for_oc_contacts(
        str(tmp_path / "DO001 ACME"), timeout=30, path=str(path)
    )

    assert result["logistics_coordinator"] == "Jane"
    assert calls["n"] >= 3


def test_wait_for_oc_contacts_gives_up_after_timeout(tmp_path, monkeypatch):
    path = tmp_path / "OC_Contacts.xlsx"
    _write_oc_contacts(path, [])
    monkeypatch.setattr(excel_sync.time, "sleep", lambda s: None)

    assert excel_sync.wait_for_oc_contacts(
        str(tmp_path / "DO001 ACME"), timeout=0, path=str(path)
    ) == {}


def test_wait_for_shipping_result_ignores_stale_generic_row(tmp_path, monkeypatch):
    """A generic row already present before the flow ran must not be adopted."""
    path = tmp_path / "Dossier_Shipping_Date.xlsx"
    _write_shipping_dates(path, [
        ["Shipping Documents and Invoices", None, "Left over from a previous order", None],
    ])
    monkeypatch.setattr(excel_sync.time, "sleep", lambda s: None)
    later = excel_sync.workbook_mtime(str(path)) + 60

    result = excel_sync.wait_for_shipping_result(
        str(tmp_path / "DO001 ACME"), since_mtime=later, timeout=0, path=str(path)
    )

    assert result == {}


def test_wait_for_shipping_result_accepts_generic_row_written_after_trigger(tmp_path, monkeypatch):
    path = tmp_path / "Dossier_Shipping_Date.xlsx"
    _write_shipping_dates(path, [
        ["Shipping Documents and Invoices", None, "No shipping documents yet", None],
    ])
    monkeypatch.setattr(excel_sync.time, "sleep", lambda s: None)

    result = excel_sync.wait_for_shipping_result(
        str(tmp_path / "DO001 ACME"), since_mtime=0.0, timeout=0, path=str(path)
    )

    assert result["reasoning"] == "No shipping documents yet"


def test_wait_for_shipping_result_prefers_own_row_regardless_of_mtime(tmp_path, monkeypatch):
    path = tmp_path / "Dossier_Shipping_Date.xlsx"
    _write_shipping_dates(path, [
        ["DO001 ACME", datetime.datetime(2026, 6, 30), "FedEx POD", "POD.pdf"],
    ])
    monkeypatch.setattr(excel_sync.time, "sleep", lambda s: None)
    later = excel_sync.workbook_mtime(str(path)) + 60

    result = excel_sync.wait_for_shipping_result(
        str(tmp_path / "DO001 ACME"), since_mtime=later, timeout=0, path=str(path)
    )

    assert result["shipping_date"] == "6/30/2026"


def test_workbook_mtime_returns_zero_for_missing_file(tmp_path):
    assert excel_sync.workbook_mtime(str(tmp_path / "nope.xlsx")) == 0.0


def test_flow_result_timeout_reads_config(monkeypatch):
    monkeypatch.setattr(
        "webapp.backend.settings.load_config", lambda: {"flow_result_timeout_seconds": 7}
    )
    assert excel_sync._flow_result_timeout() == 7.0


def test_flow_result_timeout_falls_back_on_bad_value(monkeypatch):
    monkeypatch.setattr(
        "webapp.backend.settings.load_config", lambda: {"flow_result_timeout_seconds": "abc"}
    )
    assert excel_sync._flow_result_timeout() == 90.0


# ── _excel_root (per-installation workbook folder) ───────────────────────────

def test_excel_root_uses_configured_value_when_set(monkeypatch):
    monkeypatch.setattr(
        "webapp.backend.settings.load_config",
        lambda: {"excel_root": r"C:\Users\Eric\OneDrive - ZwickRoell GmbH & Co. KG\Documents\EricProject"},
    )
    assert excel_sync._excel_root() == (
        r"C:\Users\Eric\OneDrive - ZwickRoell GmbH & Co. KG\Documents\EricProject"
    )


def test_excel_root_falls_back_to_default_when_config_blank(monkeypatch):
    monkeypatch.setattr("webapp.backend.settings.load_config", lambda: {"excel_root": ""})
    assert excel_sync._excel_root() == excel_sync._DEFAULT_ERIC_PROJECT_ROOT
