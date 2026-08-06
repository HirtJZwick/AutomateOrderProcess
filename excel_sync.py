"""
excel_sync.py
-------------
Reads contact and shipping-date data that a Power Automate flow has written to
two shared Excel workbooks, keyed by order folder name, and matches rows back
to a local order folder via `storage.folder_identity()`.

This module replaces `llm_extract_phi.py` / `llm_extract.py`: order-confirmation
contacts (logistics coordinator + RSM) and the shipping date are no longer
derived by an LLM reading PDFs directly. Instead:

  1. `ingest.py` triggers a Power Automate flow (see `power_automate.py`) for
     the order's folder.
  2. The flow reads the order's documents itself (via its own SharePoint/
     OneDrive connectors) and appends/updates a row in one of these workbooks.
  3. Once the flow's HTTP trigger responds (synchronously, HTTP 200), this
     module reads the workbook and looks up the row matching the folder that
     was just processed.

File locations are currently hard-coded to the shared OneDrive folder Eric and
Anita use; update the two path constants below if that location changes.
"""
from __future__ import annotations

import os
import time

import openpyxl

import storage

# Name of the per-order subfolder the shipping-date flow searches. The flow
# currently writes this generic name into Folder_Name instead of the parent
# order folder, so rows carrying it are treated as "belongs to whichever order
# was just processed" rather than as a different order's row.
_GENERIC_SHIPPING_FOLDER_NAMES = {
    storage.folder_identity("Shipping Documents and Invoices"),
}


class WorkbookUnavailable(RuntimeError):
    """A flow-output workbook exists but could not be opened for reading.

    Almost always means the file is currently locked — open in Excel on this
    machine, or momentarily locked while OneDrive syncs the flow's write. The
    caller should report this rather than silently treating it as "no data",
    because the row the flow just wrote is very likely in there.
    """

# Fallback default (the developer's own shared folder) used only when
# config.json has no "excel_root" set. Each installation should normally set
# its own "excel_root" in config.json to the local, OneDrive-synced folder
# holding its own two flow-output workbooks — see webapp/backend/settings.py.
_DEFAULT_ERIC_PROJECT_ROOT = (
    r"C:\Users\Hirtj\OneDrive - ZwickRoell GmbH & Co. KG\Documents\EricProject"
)


def _excel_root() -> str:
    try:
        from webapp.backend.settings import load_config

        configured = load_config().get("excel_root")
    except Exception:
        configured = None
    return configured or _DEFAULT_ERIC_PROJECT_ROOT


def _oc_contacts_path() -> str:
    return os.path.join(_excel_root(), "OC_Contacts.xlsx")


def _shipping_date_path() -> str:
    return os.path.join(_excel_root(), "Dossier_Shipping_Date.xlsx")


# Kept as module-level names for backwards compatibility with any external
# callers; resolved once at import time. Prefer _oc_contacts_path() /
# _shipping_date_path() internally so a config.json change takes effect after
# a restart without needing to re-import this module.
OC_CONTACTS_PATH = _oc_contacts_path()
SHIPPING_DATE_PATH = _shipping_date_path()

# OC_Contacts.xlsx column -> orders table column.
_OC_CONTACT_COLUMNS = {
    "Logistics_Coordinator": "logistics_coordinator",
    "Logistics_Coordinator_Email": "logistics_coordinator_email",
    "RSM": "rsm",
    "RSM_Email": "rsm_email",
}


def _load_workbook(path: str, attempts: int = 3, delay: float = 0.75):
    """Open `path` read-only, retrying briefly while it is locked.

    A OneDrive sync of the row the flow just appended locks the file for a
    moment, so a couple of quick retries turn a hard failure into a normal
    read. A lock held by an open Excel window never clears on its own, so the
    retries are deliberately short and the error is then raised.
    """
    last_error = None
    for attempt in range(attempts):
        try:
            return openpyxl.load_workbook(path, data_only=True, read_only=True)
        except PermissionError as exc:
            last_error = exc
            if attempt < attempts - 1:
                time.sleep(delay)
        except OSError as exc:  # corrupt/partial file mid-sync
            last_error = exc
            break
    raise WorkbookUnavailable(
        f"Could not read {path}: {last_error}. The file is locked — close it in "
        f"Excel (and let OneDrive finish syncing), then refresh again."
    )


def _read_rows(path: str) -> list[dict]:
    """Read all data rows of the first sheet as header->value dicts.

    Tolerant of blank leading rows/columns (as produced by the flow) — finds
    the header row by locating "Folder_Name" in the first column. Returns []
    if the file doesn't exist yet or has no recognizable header.

    Raises `WorkbookUnavailable` when the file exists but is locked.
    """
    if not os.path.exists(path):
        return []
    wb = _load_workbook(path)
    try:
        ws = wb.worksheets[0]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    header_idx = next(
        (i for i, row in enumerate(rows) if row and row[0] == "Folder_Name"), None
    )
    if header_idx is None:
        return []
    headers = rows[header_idx]
    out = []
    for row in rows[header_idx + 1:]:
        if not row or not row[0]:
            continue
        out.append({headers[i]: row[i] for i in range(len(headers)) if i < len(row)})
    return out


def _matching_rows(rows: list[dict], folder: str) -> list[dict]:
    identity = storage.folder_identity(folder)
    return [
        row for row in rows
        if storage.folder_identity(str(row.get("Folder_Name") or "")) == identity
    ]


def _format_date(value) -> str | None:
    """Format a Shipping_Date cell as 'M/D/YYYY' (matches storage._parse_order_date)."""
    if value is None:
        return None
    if hasattr(value, "strftime"):
        return f"{value.month}/{value.day}/{value.year}"
    text = str(value).strip()
    return text or None


def lookup_oc_contacts(folder: str, path: str | None = None) -> dict:
    """Return the contact fields for `folder`'s matching row, or {} if absent.

    If multiple rows match (the flow may append duplicates on repeated runs),
    the LAST matching row wins (most recent).
    """
    matches = _matching_rows(_read_rows(path or _oc_contacts_path()), folder)
    if not matches:
        return {}
    match = matches[-1]
    result = {}
    for col, field in _OC_CONTACT_COLUMNS.items():
        value = match.get(col)
        if value not in (None, ""):
            result[field] = str(value).strip()
    return result


def lookup_shipping_date(folder: str, path: str | None = None) -> dict:
    """Return {"shipping_date", "reasoning", "source_document"} for `folder`'s
    matching row, or {} if no row matches or the row has neither a date nor a
    reason to report.

    `shipping_date` may be None while `reasoning` is still populated — this
    happens when the flow couldn't determine a date but recorded why (e.g. no
    shipping documents were found yet). Callers use the reasoning in that case
    to explain to the user why no shipping date is available.
    """
    matches = _matching_rows(_read_rows(path or _shipping_date_path()), folder)
    if not matches:
        return {}
    match = matches[-1]
    shipping_date = _format_date(match.get("Shipping_Date"))
    reasoning = match.get("Reason")
    source_document = match.get("Document")
    if not shipping_date and not reasoning:
        return {}
    return {
        "shipping_date": shipping_date,
        "reasoning": reasoning,
        "source_document": source_document,
    }


def workbook_mtime(path: str | None = None) -> float:
    """Last-modified time of a workbook, or 0.0 when it does not exist yet.

    Captured *before* a flow is triggered so `wait_for_shipping_result()` can
    tell a row the flow just wrote from a stale generic row left by an earlier
    order.
    """
    try:
        return os.path.getmtime(path or _shipping_date_path())
    except OSError:
        return 0.0


def _flow_result_timeout() -> float:
    """Seconds to wait for a flow's row to reach the local synced workbook."""
    try:
        from webapp.backend.settings import load_config

        configured = load_config().get("flow_result_timeout_seconds")
    except Exception:
        configured = None
    try:
        return max(0.0, float(configured))
    except (TypeError, ValueError):
        return 90.0


def _poll(fetch, timeout: float, interval: float = 2.0) -> dict:
    """Call `fetch()` until it returns something truthy or `timeout` elapses.

    The flows respond HTTP 200 as soon as they have written their row to
    SharePoint, but that row only reaches the *local* OneDrive-synced copy of
    the workbook a few seconds later. Reading once therefore usually finds
    nothing and the result is silently lost; polling closes that gap.
    """
    deadline = time.monotonic() + timeout
    while True:
        result = fetch()
        if result:
            return result
        if time.monotonic() >= deadline:
            return {}
        time.sleep(min(interval, max(0.0, deadline - time.monotonic())))


def wait_for_oc_contacts(
    folder: str, timeout: float | None = None, path: str | None = None
) -> dict:
    """`lookup_oc_contacts()`, retried until the flow's row syncs down."""
    timeout = _flow_result_timeout() if timeout is None else timeout
    return _poll(lambda: lookup_oc_contacts(folder, path=path), timeout)


def wait_for_shipping_result(
    folder: str,
    since_mtime: float = 0.0,
    timeout: float | None = None,
    path: str | None = None,
) -> dict:
    """The shipping result for `folder`, retried until the flow's row syncs down.

    A row naming `folder` is accepted at any time. The generic
    "Shipping Documents and Invoices" row is only accepted once the workbook has
    been modified since `since_mtime`, i.e. it was written by *this* run — a
    stale generic row from a previous order must never be adopted.
    """
    timeout = _flow_result_timeout() if timeout is None else timeout
    resolved = path or _shipping_date_path()

    def fetch():
        match = lookup_shipping_date(folder, path=resolved)
        if match:
            return match
        if workbook_mtime(resolved) > since_mtime:
            return lookup_latest_shipping_result(folder, path=resolved)
        return {}

    return _poll(fetch, timeout)


def lookup_latest_shipping_result(folder: str = "", path: str | None = None) -> dict:
    """Return the latest usable row that may legitimately belong to `folder`.

    The current flow writes the generic shipping-subfolder name ("Shipping
    Documents and Invoices") into Folder_Name instead of the parent order
    folder, so immediately after the synchronous flow returns HTTP 200 the
    latest such row is the only way to retrieve the result it just produced.

    Rows naming a *different* order folder are skipped: borrowing them would
    stamp another order's shipping date and reasoning onto this order. Only
    generic rows, or rows already naming `folder`, are accepted. Scanning stops
    at the first usable row, so a stale generic row further up cannot win over
    a newer one.
    """
    own_identity = storage.folder_identity(folder) if folder else None
    for row in reversed(_read_rows(path or _shipping_date_path())):
        identity = storage.folder_identity(str(row.get("Folder_Name") or ""))
        if identity not in _GENERIC_SHIPPING_FOLDER_NAMES and identity != own_identity:
            continue  # belongs to a different order
        shipping_date = _format_date(row.get("Shipping_Date"))
        reasoning = row.get("Reason")
        if shipping_date or reasoning:
            return {
                "shipping_date": shipping_date,
                "reasoning": reasoning,
                "source_document": row.get("Document"),
            }
    return {}
