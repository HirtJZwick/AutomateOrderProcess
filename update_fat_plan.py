"""
update_fat_plan.py
------------------
Write a staged order (from eric_orders.db) into the FAT_INSTALL Plan Worksheet.

openpyxl preserves existing formulas and formatting for untouched cells. We only
set the specific cells in the MAPPING below. The workbook is matched by the
'FAT_INSTALL*.xlsx' prefix in the project folder.

Contacts: the 'Customer Contact' row (A19) is filled by parsing the technical /
shipping contact string into name / phone / email.
"""
from __future__ import annotations

import glob
import os
import re
import sys

import openpyxl

import storage

SHEET_NAME = "FAT Plan"

# Direct field -> cell mapping (header block of the FAT Plan sheet).
SCALAR_MAPPING = {
    "customer_name": "B3",       # Customer Name
    "ship_to_address": "B4",     # Install Location
}


def find_workbook(folder: str) -> str | None:
    hits = [
        p
        for p in glob.glob(os.path.join(folder, "FAT_INSTALL*.xlsx"))
        if not os.path.basename(p).startswith("~$")
        and "BACKUP" not in os.path.basename(p).upper()
    ]
    return hits[0] if hits else None


def _parse_contact(raw: str) -> dict:
    """'Porter, Katie /  +1(520)808-5695 x  katie@polychemistry.com'
    -> {name, phone, email}."""
    out = {"name": None, "phone": None, "email": None}
    if not raw:
        return out
    email = re.search(r"[\w.+-]+@[\w.-]+\.\w+", raw)
    if email:
        out["email"] = email.group(0)
        raw = raw.replace(email.group(0), "")
    phone = re.search(r"\+?\d[\d()\-\s]{6,}\d", raw)
    if phone:
        out["phone"] = phone.group(0).strip()
        raw = raw.replace(phone.group(0), "")
    name = raw.replace("x", " ").strip(" /").strip()
    out["name"] = re.sub(r"\s+", " ", name) or None
    return out


def write_plan(order: dict, xlsx_path: str) -> list[str]:
    """Apply the order to the workbook in place. Returns a list of changes."""
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
    changes: list[str] = []

    def setc(cell, value):
        if value not in (None, ""):
            ws[cell] = value
            changes.append(f"{cell} = {value!r}")

    # Document Number e.g. DO737348
    if order.get("dossier_no"):
        setc("B2", f"DO{order['dossier_no']}")

    for field, cell in SCALAR_MAPPING.items():
        setc(cell, order.get(field))

    # Customer Contact row (A19): name / company / phone / email
    contact = _parse_contact(order.get("technical_contact") or order.get("shipping_contact") or "")
    setc("B19", contact["name"])
    setc("C19", order.get("customer_name"))
    setc("D19", contact["phone"])
    setc("E19", contact["email"])

    wb.save(xlsx_path)
    return changes


def main(argv):
    folder = os.path.dirname(os.path.abspath(__file__))
    dossier = argv[1] if len(argv) > 1 else None

    conn = storage.connect()
    storage.init_db(conn)
    if dossier:
        order = storage.get_order(conn, dossier)
    else:
        cur = conn.execute("SELECT * FROM orders ORDER BY updated_at DESC LIMIT 1")
        row = cur.fetchone()
        order = dict(row) if row else None
    conn.close()

    if not order:
        print("ERROR: no staged order found to write.", file=sys.stderr)
        return 5

    xlsx = find_workbook(folder)
    if not xlsx:
        print("ERROR: FAT_INSTALL*.xlsx not found.", file=sys.stderr)
        return 6

    changes = write_plan(order, xlsx)
    print(f"Updated {os.path.basename(xlsx)} for dossier {order.get('dossier_no')}:")
    for c in changes:
        print("  " + c)
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
